import io
import csv
import json
from flask import Blueprint, render_template, request, send_file, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from reportlab.lib.pagesizes import A4, A3, letter, legal, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, PageBreak, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from utils.file_utils import make_zip

bp = Blueprint("spreadsheet", __name__)

EXCEL_ACCEPT = ".xlsx,.xlsm,.xls"
EXCEL_EXTS = {"xlsx", "xlsm", "xls"}
MAX_PDF_ROWS_PER_SHEET = 5000


# ── Reader helpers ─────────────────────────────

def _ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def read_workbook(file_data: bytes, filename: str) -> dict[str, list[list]]:
    """Read any supported Excel file into {sheet_name: [[row], ...]}."""
    ext = _ext(filename)
    if ext in ("xlsx", "xlsm"):
        wb = load_workbook(io.BytesIO(file_data), data_only=True, read_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([_normalize_cell(v) for v in row])
            sheets[name] = rows
        wb.close()
        return sheets

    if ext == "xls":
        if not HAS_XLRD:
            raise RuntimeError("Legacy .xls support requires the 'xlrd' package. Install: pip install xlrd")
        book = xlrd.open_workbook(file_contents=file_data)
        sheets = {}
        for name in book.sheet_names():
            sh = book.sheet_by_name(name)
            rows = []
            for r in range(sh.nrows):
                rows.append([_normalize_cell(sh.cell_value(r, c)) for c in range(sh.ncols)])
            sheets[name] = rows
        return sheets

    raise ValueError(f"Unsupported file type: .{ext}")


def _normalize_cell(v):
    if v is None:
        return ""
    from datetime import datetime, date, time
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    return v


def _safe_sheet_name(name: str, taken: set) -> str:
    """Return an Excel-safe, unique sheet name."""
    base = "".join(c for c in (name or "Sheet") if c not in "[]:*?/\\").strip() or "Sheet"
    base = base[:31]
    candidate = base
    i = 2
    while candidate in taken:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    taken.add(candidate)
    return candidate


# ── Page Routes ─────────────────────────────────

@bp.route("/excel-to-csv")
def excel_to_csv_page():
    return render_template("upload_tool.html",
        title="Excel to CSV / JSON",
        description="Export sheets from an Excel workbook to CSV or JSON",
        endpoint="/spreadsheet/excel-to-csv",
        accept=EXCEL_ACCEPT,
        multiple=False,
        options=[
            {"type": "select", "name": "format", "label": "Output Format",
             "choices": [
                 {"value": "csv", "label": "CSV"},
                 {"value": "json", "label": "JSON (array of objects, first row as keys)"},
                 {"value": "json_array", "label": "JSON (array of arrays)"},
             ]},
            {"type": "text", "name": "sheet", "label": "Sheet name (leave blank for all sheets)",
             "placeholder": "e.g. Sheet1"},
        ])


@bp.route("/csv-to-excel")
def csv_to_excel_page():
    return render_template("upload_tool.html",
        title="CSV / JSON to Excel",
        description="Build an Excel workbook from one or more CSV or JSON files (one sheet per file)",
        endpoint="/spreadsheet/csv-to-excel",
        accept=".csv,.json",
        multiple=True,
        options=[
            {"type": "checkbox", "name": "bold_headers", "label": "Formatting",
             "check_label": "Bold and shade the header row", "default": True},
        ])


@bp.route("/excel-to-pdf")
def excel_to_pdf_page():
    return render_template("upload_tool.html",
        title="Excel to PDF",
        description="Convert an Excel workbook to PDF (one section per sheet). Basic table rendering — not pixel-perfect.",
        endpoint="/spreadsheet/excel-to-pdf",
        accept=EXCEL_ACCEPT,
        multiple=False,
        options=[
            {"type": "select", "name": "size", "label": "Page Size",
             "choices": [
                 {"value": "A4", "label": "A4"},
                 {"value": "letter", "label": "Letter"},
                 {"value": "A3", "label": "A3 (wide tables)"},
                 {"value": "legal", "label": "Legal"},
             ]},
            {"type": "select", "name": "orientation", "label": "Orientation",
             "choices": [
                 {"value": "landscape", "label": "Landscape"},
                 {"value": "portrait", "label": "Portrait"},
             ]},
            {"type": "number", "name": "fontsize", "label": "Font Size",
             "default": 8, "min": 5, "max": 14},
        ])


@bp.route("/merge")
def merge_page():
    return render_template("upload_tool.html",
        title="Merge Workbooks",
        description="Combine multiple Excel files into a single workbook. Each source sheet becomes one sheet in the output.",
        endpoint="/spreadsheet/merge",
        accept=EXCEL_ACCEPT,
        multiple=True,
        options=[
            {"type": "checkbox", "name": "prefix", "label": "Sheet names",
             "check_label": "Prefix each sheet with its source filename", "default": True},
        ])


@bp.route("/split")
def split_page():
    return render_template("upload_tool.html",
        title="Split Sheets",
        description="Export every sheet in a workbook as its own .xlsx file (bundled as a ZIP).",
        endpoint="/spreadsheet/split",
        accept=EXCEL_ACCEPT,
        multiple=False,
        options=[])


@bp.route("/info")
def info_page():
    return render_template("upload_tool.html",
        title="Excel Info & Preview",
        description="Inspect sheet names, row/column counts, and preview the first rows of each sheet.",
        endpoint="/spreadsheet/info",
        accept=EXCEL_ACCEPT,
        multiple=False,
        options=[
            {"type": "number", "name": "preview_rows", "label": "Preview rows per sheet",
             "default": 10, "min": 0, "max": 100},
        ])


@bp.route("/csv-tools")
def csv_tools_page():
    return render_template("upload_tool.html",
        title="CSV Toolkit",
        description="Sort, filter, or de-duplicate rows in a CSV file",
        notes=(
            "<p>Operations run in order: <strong>filter → dedupe → sort</strong>. "
            "Use <code>column_name</code> or a 1-based column number. "
            "Filter operators: <code>=</code>, <code>!=</code>, <code>contains</code>, "
            "<code>startswith</code>, <code>endswith</code>, <code>&gt;</code>, "
            "<code>&gt;=</code>, <code>&lt;</code>, <code>&lt;=</code>, "
            "<code>empty</code>, <code>notempty</code>.</p>"
        ),
        endpoint="/spreadsheet/csv-tools",
        accept=".csv,.tsv,.txt",
        multiple=False,
        options=[
            {"type": "checkbox", "name": "has_header", "label": "Header row",
             "default": True, "check_label": "First row is a header"},
            {"type": "select", "name": "delimiter", "label": "Delimiter", "default": "auto",
             "choices": [
                 {"value": "auto",  "label": "Auto-detect"},
                 {"value": ",",     "label": "Comma"},
                 {"value": ";",     "label": "Semicolon"},
                 {"value": "\\t",   "label": "Tab"},
                 {"value": "|",     "label": "Pipe"},
             ]},
            {"type": "text", "name": "filter", "label": "Filter (optional)",
             "placeholder": "e.g. status = active    |    age >= 18    |    city contains London"},
            {"type": "text", "name": "sort_by", "label": "Sort by (optional)",
             "placeholder": "column_name or 1-based column number"},
            {"type": "select", "name": "sort_dir", "label": "Sort direction", "default": "asc",
             "choices": [
                 {"value": "asc", "label": "Ascending"},
                 {"value": "desc", "label": "Descending"},
             ]},
            {"type": "select", "name": "sort_type", "label": "Sort as", "default": "smart",
             "choices": [
                 {"value": "smart",  "label": "Smart (number if numeric, else text)"},
                 {"value": "text",   "label": "Text"},
                 {"value": "number", "label": "Number"},
             ]},
            {"type": "select", "name": "dedupe", "label": "Remove duplicates", "default": "none",
             "choices": [
                 {"value": "none",   "label": "Keep all rows"},
                 {"value": "full",   "label": "On full row"},
                 {"value": "bycol",  "label": "On selected columns (below)"},
             ]},
            {"type": "text", "name": "dedupe_cols", "label": "Dedupe columns (comma-separated)",
             "placeholder": "e.g. email, phone"},
        ])


# ── Processing Routes ───────────────────────────

@bp.route("/excel-to-csv", methods=["POST"])
def excel_to_csv():
    files = request.files.getlist("files")
    if not files or not files[0].filename:
        return jsonify(error="No file uploaded."), 400

    fmt = request.form.get("format", "csv")
    target_sheet = (request.form.get("sheet") or "").strip()

    try:
        sheets = read_workbook(files[0].read(), files[0].filename)
    except Exception as e:
        return jsonify(error=f"Could not read workbook: {e}"), 400

    if target_sheet:
        if target_sheet not in sheets:
            return jsonify(error=f"Sheet '{target_sheet}' not found. Available: {', '.join(sheets.keys())}"), 400
        sheets = {target_sheet: sheets[target_sheet]}

    outputs = []
    for name, rows in sheets.items():
        if fmt == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf, lineterminator="\n")
            for row in rows:
                writer.writerow(row)
            outputs.append((f"{name}.csv", buf.getvalue().encode("utf-8-sig")))
        elif fmt == "json":
            data = _rows_to_dicts(rows)
            outputs.append((f"{name}.json", json.dumps(data, ensure_ascii=False, indent=2, default=str).encode("utf-8")))
        else:  # json_array
            outputs.append((f"{name}.json", json.dumps(rows, ensure_ascii=False, indent=2, default=str).encode("utf-8")))

    base = files[0].filename.rsplit(".", 1)[0]

    if len(outputs) == 1:
        fname, data = outputs[0]
        mime = "text/csv" if fname.endswith(".csv") else "application/json"
        return send_file(io.BytesIO(data), mimetype=mime,
                         as_attachment=True, download_name=fname)

    zip_buf = make_zip(outputs)
    return send_file(zip_buf, mimetype="application/zip",
                     as_attachment=True, download_name=f"{base}_{fmt}.zip")


def _rows_to_dicts(rows):
    if not rows:
        return []
    headers = [str(h) if h != "" else f"col_{i+1}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, r)) for r in rows[1:]]


@bp.route("/csv-to-excel", methods=["POST"])
def csv_to_excel():
    files = request.files.getlist("files")
    if not files or not files[0].filename:
        return jsonify(error="No file uploaded."), 400

    bold_headers = request.form.get("bold_headers") == "on"

    wb = Workbook()
    wb.remove(wb.active)
    taken_names = set()

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E8EBF0", end_color="E8EBF0", fill_type="solid")

    for f in files:
        name = f.filename
        ext = _ext(name)
        data = f.read()
        base = name.rsplit(".", 1)[0]
        sheet_name = _safe_sheet_name(base, taken_names)
        ws = wb.create_sheet(title=sheet_name)

        try:
            if ext == "csv":
                text = data.decode("utf-8-sig", errors="replace")
                reader = csv.reader(io.StringIO(text))
                rows = list(reader)
            elif ext == "json":
                parsed = json.loads(data.decode("utf-8", errors="replace"))
                rows = _json_to_rows(parsed)
            else:
                return jsonify(error=f"Unsupported file type: {name}"), 400
        except Exception as e:
            return jsonify(error=f"Failed to parse {name}: {e}"), 400

        for row in rows:
            ws.append([_coerce(v) for v in row])

        if bold_headers and rows:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left")

        _autosize_columns(ws, rows)

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    out_name = (files[0].filename.rsplit(".", 1)[0] if len(files) == 1 else "workbook") + ".xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=out_name)


def _json_to_rows(parsed):
    if isinstance(parsed, list):
        if not parsed:
            return []
        if isinstance(parsed[0], dict):
            headers = list({k: None for row in parsed for k in row.keys()}.keys())
            rows = [headers]
            for obj in parsed:
                rows.append([obj.get(h, "") for h in headers])
            return rows
        if isinstance(parsed[0], list):
            return parsed
    raise ValueError("JSON must be an array of objects or an array of arrays")


def _coerce(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    if isinstance(v, str):
        s = v.strip()
        if s.lstrip("-").isdigit() and len(s) < 16:
            try:
                return int(s)
            except ValueError:
                pass
        try:
            if "." in s or "e" in s.lower():
                return float(s)
        except ValueError:
            pass
    return v


def _autosize_columns(ws, rows, max_width=60):
    if not rows:
        return
    col_count = max(len(r) for r in rows)
    for col in range(col_count):
        max_len = 0
        for r in rows:
            if col < len(r):
                val = r[col]
                if val is None:
                    continue
                length = len(str(val))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[get_column_letter(col + 1)].width = min(max(max_len + 2, 10), max_width)


@bp.route("/excel-to-pdf", methods=["POST"])
def excel_to_pdf():
    files = request.files.getlist("files")
    if not files or not files[0].filename:
        return jsonify(error="No file uploaded."), 400

    size_name = request.form.get("size", "A4")
    orientation = request.form.get("orientation", "landscape")
    fontsize = int(request.form.get("fontsize", 8))

    page_size_map = {"A4": A4, "A3": A3, "letter": letter, "legal": legal}
    page_size = page_size_map.get(size_name, A4)
    if orientation == "landscape":
        page_size = landscape(page_size)

    try:
        sheets = read_workbook(files[0].read(), files[0].filename)
    except Exception as e:
        return jsonify(error=f"Could not read workbook: {e}"), 400

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=page_size,
                            leftMargin=0.4 * inch, rightMargin=0.4 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading2"],
                                 fontName="Helvetica-Bold", fontSize=fontsize + 4,
                                 spaceAfter=8)

    story = []
    sheet_items = list(sheets.items())
    for i, (name, rows) in enumerate(sheet_items):
        if i > 0:
            story.append(PageBreak())
        story.append(Paragraph(name, title_style))

        if not rows:
            story.append(Paragraph("<i>(empty sheet)</i>", styles["Normal"]))
            continue

        truncated = len(rows) > MAX_PDF_ROWS_PER_SHEET
        data_rows = rows[:MAX_PDF_ROWS_PER_SHEET]
        max_cols = max((len(r) for r in data_rows), default=0)
        normalized = [[_pdf_cell(c) for c in r] + [""] * (max_cols - len(r)) for r in data_rows]

        if not normalized or max_cols == 0:
            continue

        tbl = Table(normalized, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), fontsize),
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.90, 0.92, 0.96)),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(tbl)

        if truncated:
            story.append(Spacer(1, 6))
            story.append(Paragraph(
                f"<i>Sheet truncated at {MAX_PDF_ROWS_PER_SHEET:,} rows (of {len(rows):,}).</i>",
                styles["Normal"]))

    if not story:
        story.append(Paragraph("(workbook is empty)", styles["Normal"]))

    try:
        pdf.build(story)
    except Exception as e:
        return jsonify(error=f"PDF layout failed (table too wide?). Try a larger page size or smaller font. Details: {str(e)[:150]}"), 400

    buf.seek(0)
    base = files[0].filename.rsplit(".", 1)[0]
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=f"{base}.pdf")


def _pdf_cell(v):
    if v is None:
        return ""
    s = str(v)
    if len(s) > 200:
        s = s[:197] + "..."
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


@bp.route("/merge", methods=["POST"])
def merge():
    files = request.files.getlist("files")
    if not files or not files[0].filename:
        return jsonify(error="No files uploaded."), 400

    use_prefix = request.form.get("prefix") == "on"

    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    taken = set()

    for f in files:
        data = f.read()
        source = f.filename.rsplit(".", 1)[0]
        try:
            sheets = read_workbook(data, f.filename)
        except Exception as e:
            return jsonify(error=f"Failed to read {f.filename}: {e}"), 400

        for sheet_name, rows in sheets.items():
            desired = f"{source}_{sheet_name}" if use_prefix else sheet_name
            name = _safe_sheet_name(desired, taken)
            ws = out_wb.create_sheet(title=name)
            for row in rows:
                ws.append([_coerce_keep(v) for v in row])

    if not out_wb.sheetnames:
        out_wb.create_sheet("Sheet1")

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="merged.xlsx")


def _coerce_keep(v):
    if v is None:
        return ""
    return v


@bp.route("/split", methods=["POST"])
def split():
    files = request.files.getlist("files")
    if not files or not files[0].filename:
        return jsonify(error="No file uploaded."), 400

    try:
        sheets = read_workbook(files[0].read(), files[0].filename)
    except Exception as e:
        return jsonify(error=f"Could not read workbook: {e}"), 400

    if not sheets:
        return jsonify(error="Workbook contains no sheets."), 400

    base = files[0].filename.rsplit(".", 1)[0]
    outputs = []
    for name, rows in sheets.items():
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title=_safe_sheet_name(name, set()))
        for row in rows:
            ws.append([_coerce_keep(v) for v in row])
        buf = io.BytesIO()
        wb.save(buf)
        outputs.append((_safe_filename(name) + ".xlsx", buf.getvalue()))

    if len(outputs) == 1:
        fname, data = outputs[0]
        return send_file(io.BytesIO(data),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=fname)

    zip_buf = make_zip(outputs)
    return send_file(zip_buf, mimetype="application/zip",
                     as_attachment=True, download_name=f"{base}_sheets.zip")


def _safe_filename(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_ " else "_" for c in name).strip() or "Sheet"


@bp.route("/info", methods=["POST"])
def info():
    files = request.files.getlist("files")
    if not files or not files[0].filename:
        return jsonify(error="No file uploaded."), 400

    preview = int(request.form.get("preview_rows", 10))

    try:
        sheets = read_workbook(files[0].read(), files[0].filename)
    except Exception as e:
        return jsonify(error=f"Could not read workbook: {e}"), 400

    lines = [f"File: {files[0].filename}", f"Sheets: {len(sheets)}", ""]
    for name, rows in sheets.items():
        cols = max((len(r) for r in rows), default=0)
        lines.append(f"═ {name}")
        lines.append(f"  rows: {len(rows)}   columns: {cols}")
        if preview > 0 and rows:
            lines.append(f"  preview (first {min(preview, len(rows))} rows):")
            for r in rows[:preview]:
                cells = [str(v)[:30] for v in r]
                lines.append("    " + " | ".join(cells))
        lines.append("")

    return jsonify(text="\n".join(lines).rstrip())


@bp.route("/csv-tools", methods=["POST"])
def csv_tools():
    import re

    files = request.files.getlist("files")
    if not files or not files[0].filename:
        return jsonify(error="No file uploaded."), 400

    raw = files[0].read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("utf-8", errors="replace")

    # Delimiter
    delim = request.form.get("delimiter", "auto")
    if delim == "\\t":
        delim = "\t"
    if delim == "auto":
        try:
            delim = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|").delimiter
        except csv.Error:
            delim = ","

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = [row for row in reader]
    if not rows:
        return jsonify(error="CSV appears to be empty."), 400

    has_header = request.form.get("has_header") == "on"
    header = rows[0] if has_header else [f"col{i + 1}" for i in range(len(rows[0]))]
    data = rows[1:] if has_header else rows
    ncols = len(header)

    def _col_idx(ref: str) -> int:
        ref = ref.strip()
        if not ref:
            return -1
        # Try 1-based numeric
        if ref.isdigit():
            n = int(ref) - 1
            if 0 <= n < ncols:
                return n
        # Case-insensitive header match
        for i, h in enumerate(header):
            if h.strip().lower() == ref.lower():
                return i
        return -1

    def _as_num(s: str):
        try:
            return float(str(s).replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    # ── Filter ───────────────────────────────────
    filter_expr = request.form.get("filter", "").strip()
    if filter_expr:
        # Parse "col OP value" with OP in our set
        ops = ["!=", ">=", "<=", "=", ">", "<", "contains", "startswith", "endswith", "empty", "notempty"]
        # Try binary ops first (longest match)
        matched = None
        for op in ops:
            if op in ("empty", "notempty"):
                m = re.match(rf"^\s*(.+?)\s+{op}\s*$", filter_expr, re.IGNORECASE)
                if m:
                    matched = (m.group(1), op, "")
                    break
            else:
                m = re.match(rf"^\s*(.+?)\s*{re.escape(op)}\s*(.*)$", filter_expr)
                if m:
                    matched = (m.group(1), op, m.group(2))
                    break
        if not matched:
            return jsonify(error=f"Could not parse filter '{filter_expr}'."), 400

        col_ref, op, value = matched
        ci = _col_idx(col_ref)
        if ci < 0:
            return jsonify(error=f"Filter column '{col_ref}' not found."), 400

        def _keep(row):
            if ci >= len(row):
                cell = ""
            else:
                cell = str(row[ci])
            if op == "empty":
                return cell.strip() == ""
            if op == "notempty":
                return cell.strip() != ""
            if op == "contains":
                return value.lower() in cell.lower()
            if op == "startswith":
                return cell.lower().startswith(value.lower())
            if op == "endswith":
                return cell.lower().endswith(value.lower())
            if op == "=":
                return cell == value
            if op == "!=":
                return cell != value
            # Numeric comparisons
            cn, vn = _as_num(cell), _as_num(value)
            if cn is None or vn is None:
                return False
            if op == ">": return cn > vn
            if op == ">=": return cn >= vn
            if op == "<": return cn < vn
            if op == "<=": return cn <= vn
            return False

        data = [r for r in data if _keep(r)]

    # ── Dedupe ───────────────────────────────────
    dedupe = request.form.get("dedupe", "none")
    if dedupe == "full":
        seen, unique = set(), []
        for r in data:
            key = tuple(r)
            if key not in seen:
                seen.add(key); unique.append(r)
        data = unique
    elif dedupe == "bycol":
        col_refs = [c.strip() for c in request.form.get("dedupe_cols", "").split(",") if c.strip()]
        indices = [_col_idx(c) for c in col_refs]
        if any(i < 0 for i in indices):
            bad = [c for c, i in zip(col_refs, indices) if i < 0]
            return jsonify(error=f"Dedupe column(s) not found: {', '.join(bad)}"), 400
        if not indices:
            return jsonify(error="Please list at least one dedupe column."), 400
        seen, unique = set(), []
        for r in data:
            key = tuple(r[i] if i < len(r) else "" for i in indices)
            if key not in seen:
                seen.add(key); unique.append(r)
        data = unique

    # ── Sort ─────────────────────────────────────
    sort_by = request.form.get("sort_by", "").strip()
    if sort_by:
        si = _col_idx(sort_by)
        if si < 0:
            return jsonify(error=f"Sort column '{sort_by}' not found."), 400
        sort_type = request.form.get("sort_type", "smart")
        desc = request.form.get("sort_dir", "asc") == "desc"

        def _key(row):
            v = row[si] if si < len(row) else ""
            if sort_type == "text":
                return (0, str(v))
            n = _as_num(v)
            if sort_type == "number":
                return (0 if n is not None else 1, n if n is not None else 0)
            # smart: numbers first, lexicographic fallback
            if n is not None:
                return (0, n)
            return (1, str(v).lower())

        data.sort(key=_key, reverse=desc)

    # ── Write back ───────────────────────────────
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delim, lineterminator="\n")
    if has_header:
        writer.writerow(header)
    for r in data:
        writer.writerow(r)

    out_bytes = output.getvalue().encode("utf-8-sig")
    name = files[0].filename.rsplit(".", 1)[0] + "_processed.csv"
    return send_file(io.BytesIO(out_bytes), mimetype="text/csv",
                     as_attachment=True, download_name=name)
